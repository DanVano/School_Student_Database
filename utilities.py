import csv
import json

import pandas as pd

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


#Import the student information from a CSV file
def gather_student_info():
    with open('schools.csv', newline='') as csvfile:
        reader = csv.DictReader(csvfile)
        data = list(reader)
    schools_db = {row['ID']: {'Name': row['Name'], 'Birthday': row['Birthday'], 'Classes': json.loads(row['Classes'])} for row in data}
    return schools_db

# Save the modified schools_db to an excel file
def save_excel(schools_db):
    schools_db = dict(sorted(schools_db.items()))
    flat_data = {}
    # Save the excel file
    for id, info in schools_db.items():
        flat_data[id] = {**{'Name': info['Name'], 'Birthday': info['Birthday']}, **info['Classes']}
    df = pd.DataFrame(flat_data).T
    df.to_excel("Student_Info_Database.xlsx")
    book = load_workbook("Student_Info_Database.xlsx")
    sheet = book.active
    # Fix the formatting of the excel file for better readability
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
    # Save the workbook
    book.save("Student_Info_Database.xlsx")

